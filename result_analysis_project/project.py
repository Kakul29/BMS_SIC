"""
Result Analysis System
======================
1. Generates question_bank.xlsx  — question metadata + correct answers
2. Generates student_responses.xlsx — simulated student answer sheets
3. Analyses results per student
4. Produces one PDF per student  +  one examiner summary PDF
"""

import random
import os
import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import cm
from reportlab.platypus import (
    SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle,
    HRFlowable, PageBreak
)
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT

# ─────────────────────────────────────────────
# CONFIGURATION
# ─────────────────────────────────────────────
SUBJECTS   = ["OOP", "DBMS", "DSA", "Pandas"]
LEVELS     = ["Easy", "Intermediate", "Complex"]
NUM_QUESTIONS = 20          # total questions
NUM_STUDENTS  = 5

random.seed(42)

GRADE_TABLE = [
    (90, "O  – Outstanding"),
    (80, "A+ – Excellent"),
    (70, "A  – Very Good"),
    (60, "B+ – Good"),
    (50, "B  – Average"),
    (40, "C  – Pass"),
    (0,  "F  – Fail"),
]

OUT_DIR = "output"
os.makedirs(OUT_DIR, exist_ok=True)


# ─────────────────────────────────────────────
# HELPER – grade from percentage
# ─────────────────────────────────────────────
def get_grade(pct):
    for threshold, label in GRADE_TABLE:
        if pct >= threshold:
            return label
    return "F – Fail"


# ─────────────────────────────────────────────
# STEP 1 – Build Question Bank Excel
# ─────────────────────────────────────────────
def build_question_bank():
    """
    Columns:
        Q_No | Question | Type (MCSR/MCMR) | Subject | Level |
        Option_A | Option_B | Option_C | Option_D |
        Correct_Answer | Max_Marks
    """
    rows = []
    q_per_subject = NUM_QUESTIONS // len(SUBJECTS)
    q_no = 1

    for subj in SUBJECTS:
        for _ in range(q_per_subject):
            q_type  = random.choice(["MCSR", "MCMR"])
            level   = random.choice(LEVELS)
            options = ["Option A", "Option B", "Option C", "Option D"]

            if q_type == "MCSR":
                correct = random.choice(["A", "B", "C", "D"])
            else:
                choices = random.sample(["A", "B", "C", "D"], random.randint(2, 3))
                correct = ",".join(sorted(choices))

            max_marks = 1 if level == "Easy" else (2 if level == "Intermediate" else 3)

            rows.append({
                "Q_No":           q_no,
                "Question":       f"Q{q_no}: Sample {subj} question ({level})?",
                "Type":           q_type,
                "Subject":        subj,
                "Level":          level,
                "Option_A":       f"Answer choice A for Q{q_no}",
                "Option_B":       f"Answer choice B for Q{q_no}",
                "Option_C":       f"Answer choice C for Q{q_no}",
                "Option_D":       f"Answer choice D for Q{q_no}",
                "Correct_Answer": correct,
                "Max_Marks":      max_marks,
            })
            q_no += 1

    df = pd.DataFrame(rows)
    path = os.path.join(OUT_DIR, "question_bank.xlsx")
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Questions")
        ws = writer.sheets["Questions"]
        _style_question_sheet(ws)
    print(f"[✓] question_bank.xlsx  ({len(df)} questions)")
    return df


def _style_question_sheet(ws):
    header_fill = PatternFill("solid", fgColor="1F4E79")
    for cell in ws[1]:
        cell.fill   = header_fill
        cell.font   = Font(bold=True, color="FFFFFF", size=11)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    alt_fill = PatternFill("solid", fgColor="DEEAF1")
    for i, row in enumerate(ws.iter_rows(min_row=2), start=2):
        fill = alt_fill if i % 2 == 0 else PatternFill("solid", fgColor="FFFFFF")
        for cell in row:
            cell.fill = fill
            cell.alignment = Alignment(wrap_text=True, vertical="center")

    col_widths = [6, 40, 6, 10, 13, 22, 22, 22, 22, 16, 10]
    for col, width in zip(ws.columns, col_widths):
        ws.column_dimensions[col[0].column_letter].width = width
    ws.row_dimensions[1].height = 28


# ─────────────────────────────────────────────
# STEP 2 – Build Student Responses Excel
# ─────────────────────────────────────────────
def build_student_responses(qdf):
    """
    One row per student.
    Columns: Student_ID | Name | Total_Time_sec |
             Q1_Answer | Q1_Time | Q2_Answer | Q2_Time | …
    """
    student_names = ["Aarav Shah", "Priya Nair", "Rohan Mehta",
                     "Sneha Iyer", "Kiran Bose"][:NUM_STUDENTS]

    rows = []
    for sid, name in enumerate(student_names, start=1):
        row = {"Student_ID": f"S{sid:03d}", "Name": name}
        total_time = 0
        for _, q in qdf.iterrows():
            q_no    = q["Q_No"]
            correct = q["Correct_Answer"]
            q_type  = q["Type"]

            # Simulate: students answer correctly ~65–85 % of the time
            correct_prob = random.uniform(0.55, 0.90)
            if random.random() < correct_prob:
                answer = correct
            else:
                if q_type == "MCSR":
                    wrong_choices = [c for c in ["A","B","C","D"] if c != correct]
                    answer = random.choice(wrong_choices)
                else:
                    all_combos = []
                    for size in [1, 2, 3]:
                        import itertools
                        all_combos += [",".join(sorted(c))
                                       for c in itertools.combinations("ABCD", size)]
                    all_combos = [c for c in all_combos if c != correct]
                    answer = random.choice(all_combos)

            time_sec = random.randint(15, 180)
            total_time += time_sec
            row[f"Q{q_no}_Answer"] = answer
            row[f"Q{q_no}_Time"]   = time_sec

        row["Total_Time_sec"] = total_time
        rows.append(row)

    df = pd.DataFrame(rows)
    path = os.path.join(OUT_DIR, "student_responses.xlsx")
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Responses")
        ws = writer.sheets["Responses"]
        _style_response_sheet(ws)
    print(f"[✓] student_responses.xlsx  ({len(df)} students)")
    return df


def _style_response_sheet(ws):
    header_fill = PatternFill("solid", fgColor="375623")
    for cell in ws[1]:
        cell.fill  = header_fill
        cell.font  = Font(bold=True, color="FFFFFF", size=10)
        cell.alignment = Alignment(horizontal="center", wrap_text=True)

    for i, row in enumerate(ws.iter_rows(min_row=2), start=2):
        fill = PatternFill("solid", fgColor="E2EFDA" if i % 2 == 0 else "FFFFFF")
        for cell in row:
            cell.fill = fill
            cell.alignment = Alignment(horizontal="center", vertical="center")

    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = 12
    ws.column_dimensions["A"].width = 10
    ws.column_dimensions["B"].width = 18


# ─────────────────────────────────────────────
# STEP 3 – Analyse one student
# ─────────────────────────────────────────────
def analyse_student(student_row, qdf):
    sid   = student_row["Student_ID"]
    name  = student_row["Name"]
    total_time = student_row["Total_Time_sec"]

    records = []
    for _, q in qdf.iterrows():
        q_no    = q["Q_No"]
        correct = q["Correct_Answer"]
        given   = str(student_row.get(f"Q{q_no}_Answer", "")).strip()
        time_s  = int(student_row.get(f"Q{q_no}_Time", 0))

        if q["Type"] == "MCSR":
            is_correct = (given.strip().upper() == correct.strip().upper())
            marks = q["Max_Marks"] if is_correct else 0
        else:
            given_set   = set(x.strip().upper() for x in given.split(",") if x.strip())
            correct_set = set(x.strip().upper() for x in correct.split(",") if x.strip())
            if given_set == correct_set:
                marks      = q["Max_Marks"]
                is_correct = True
            elif given_set & correct_set:          # partial
                marks      = round(q["Max_Marks"] * len(given_set & correct_set) / len(correct_set), 1)
                is_correct = False
            else:
                marks      = 0
                is_correct = False

        records.append({
            "Q_No":      q_no,
            "Subject":   q["Subject"],
            "Level":     q["Level"],
            "Type":      q["Type"],
            "MaxMarks":  q["Max_Marks"],
            "Marks":     marks,
            "Correct":   is_correct,
            "Time":      time_s,
        })

    df = pd.DataFrame(records)

    # ── Aggregates ──────────────────────────────
    total_q         = len(df)
    correct_count   = int(df["Correct"].sum())
    total_marks     = df["MaxMarks"].sum()
    scored_marks    = df["Marks"].sum()
    percentage      = round(scored_marks / total_marks * 100, 2)
    grade           = get_grade(percentage)

    # Time stats
    max_row = df.loc[df["Time"].idxmax()]
    min_row = df.loc[df["Time"].idxmin()]
    avg_time = round(df["Time"].mean(), 1)

    # Level breakdowns
    level_stats = {}
    for lvl in LEVELS:
        sub = df[df["Level"] == lvl]
        if len(sub) == 0:
            level_stats[lvl] = {"avg_time": 0, "pct": 0}
        else:
            pct = round(sub["Marks"].sum() / sub["MaxMarks"].sum() * 100, 2)
            level_stats[lvl] = {"avg_time": round(sub["Time"].mean(), 1), "pct": pct}

    # Subject breakdowns
    subj_stats = {}
    for subj in SUBJECTS:
        sub = df[df["Subject"] == subj]
        if len(sub) == 0:
            subj_stats[subj] = 0
        else:
            subj_stats[subj] = round(sub["Marks"].sum() / sub["MaxMarks"].sum() * 100, 2)

    best_subj  = max(subj_stats, key=subj_stats.get)
    worst_subj = min(subj_stats, key=subj_stats.get)

    return {
        "Student_ID":     sid,
        "Name":           name,
        "Total_Q":        total_q,
        "Correct":        correct_count,
        "TotalMarks":     total_marks,
        "ScoredMarks":    scored_marks,
        "Percentage":     percentage,
        "Grade":          grade,
        "TotalTime":      total_time,
        "MaxTimeQ":       int(max_row["Q_No"]),
        "MaxTimeSec":     int(max_row["Time"]),
        "MinTimeQ":       int(min_row["Q_No"]),
        "MinTimeSec":     int(min_row["Time"]),
        "AvgTimeSec":     avg_time,
        "LevelStats":     level_stats,
        "SubjectStats":   subj_stats,
        "BestSubject":    best_subj,
        "WorstSubject":   worst_subj,
        "Records":        df,
        "Rank": 0,
    }


# ─────────────────────────────────────────────
# STEP 4a – Student PDF
# ─────────────────────────────────────────────
BRAND_BLUE  = colors.HexColor("#1F4E79")
BRAND_LIGHT = colors.HexColor("#DEEAF1")
ACCENT      = colors.HexColor("#2E75B6")
GREEN       = colors.HexColor("#375623")
RED         = colors.HexColor("#C00000")
GOLD        = colors.HexColor("#BF8F00")

def _styles():
    s = getSampleStyleSheet()
    s.add(ParagraphStyle("H1", parent=s["Heading1"], fontSize=18,
                         textColor=BRAND_BLUE, spaceAfter=4, alignment=TA_CENTER))
    s.add(ParagraphStyle("H2", parent=s["Heading2"], fontSize=13,
                         textColor=ACCENT, spaceBefore=10, spaceAfter=4))
    s.add(ParagraphStyle("Sub", parent=s["Normal"], fontSize=9,
                         textColor=colors.grey, alignment=TA_CENTER, spaceAfter=10))
    s.add(ParagraphStyle("Body", parent=s["Normal"], fontSize=10, spaceAfter=3))
    s.add(ParagraphStyle("Bold", parent=s["Normal"], fontSize=10,
                         fontName="Helvetica-Bold"))
    s.add(ParagraphStyle("Center", parent=s["Normal"], fontSize=10,
                         alignment=TA_CENTER))
    s.add(ParagraphStyle("GradeStyle", parent=s["Normal"], fontSize=28,
                         fontName="Helvetica-Bold", alignment=TA_CENTER,
                         textColor=BRAND_BLUE))
    return s


def _tbl_style(header_color=BRAND_BLUE):
    return TableStyle([
        ("BACKGROUND",  (0, 0), (-1,  0), header_color),
        ("TEXTCOLOR",   (0, 0), (-1,  0), colors.white),
        ("FONTNAME",    (0, 0), (-1,  0), "Helvetica-Bold"),
        ("FONTSIZE",    (0, 0), (-1,  0), 10),
        ("ALIGN",       (0, 0), (-1, -1), "CENTER"),
        ("VALIGN",      (0, 0), (-1, -1), "MIDDLE"),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, BRAND_LIGHT]),
        ("GRID",        (0, 0), (-1, -1), 0.4, colors.HexColor("#AAAAAA")),
        ("FONTSIZE",    (0, 1), (-1, -1), 9),
        ("TOPPADDING",  (0, 0), (-1, -1), 5),
        ("BOTTOMPADDING",(0, 0),(-1, -1), 5),
    ])


def generate_student_pdf(result, output_dir):
    s   = _styles()
    sid = result["Student_ID"]
    path = os.path.join(output_dir, f"{sid}_{result['Name'].replace(' ','_')}_Report.pdf")
    doc = SimpleDocTemplate(path, pagesize=A4,
                            leftMargin=1.8*cm, rightMargin=1.8*cm,
                            topMargin=1.5*cm, bottomMargin=1.5*cm)
    story = []

    # ── Header ──────────────────────────────────
    story.append(Paragraph("📋 Student Result Analysis", s["H1"]))
    story.append(Paragraph(f"{result['Name']}  |  ID: {sid}", s["Sub"]))
    story.append(HRFlowable(width="100%", thickness=2, color=BRAND_BLUE))
    story.append(Spacer(1, 0.3*cm))

    # ── Grade badge ─────────────────────────────
    pct = result["Percentage"]
    grade_color = GREEN if pct >= 60 else RED
    story.append(Paragraph(f"{pct}%", s["GradeStyle"]))
    story.append(Paragraph(
        f'<font color="#{grade_color.hexval()[2:]}"><b>{result["Grade"]}</b></font>',
        ParagraphStyle("GC", parent=s["Center"], fontSize=14)))
    story.append(Spacer(1, 0.4*cm))

    # ── Overview table ───────────────────────────
    story.append(Paragraph("Overview", s["H2"]))
    mins, secs = divmod(result["TotalTime"], 60)
    overview_data = [
        ["Metric", "Value"],
        ["Total Questions",           str(result["Total_Q"])],
        ["Correct Answers",           str(result.get("Rank", "-"))],
        ["Marks Scored / Total",      f"{result['ScoredMarks']} / {result['TotalMarks']}"],
        ["Percentage",                f"{pct}%"],
        ["Grade",                     result["Grade"]],
        ["Total Time Taken",          f"{mins}m {secs}s"],
    ]
    tbl = Table(overview_data, colWidths=[9*cm, 9*cm])
    tbl.setStyle(_tbl_style())
    story.append(tbl)
    story.append(Spacer(1, 0.4*cm))

    # ── Time Analysis ────────────────────────────
    story.append(Paragraph("Time Analysis", s["H2"]))
    time_data = [
        ["Metric", "Question #", "Time (sec)"],
        ["Longest time spent",  f"Q{result['MaxTimeQ']}", str(result["MaxTimeSec"])],
        ["Shortest time spent", f"Q{result['MinTimeQ']}", str(result["MinTimeSec"])],
        ["Average time / question", "–",               str(result["AvgTimeSec"])],
    ]
    tbl2 = Table(time_data, colWidths=[7*cm, 5*cm, 6*cm])
    tbl2.setStyle(_tbl_style(ACCENT))
    story.append(tbl2)
    story.append(Spacer(1, 0.4*cm))

    # ── Level Analysis ───────────────────────────
    story.append(Paragraph("Performance by Difficulty Level", s["H2"]))
    lvl_data = [["Level", "Avg Time (sec)", "Score %"]]
    for lvl in LEVELS:
        st = result["LevelStats"][lvl]
        lvl_data.append([lvl, str(st["avg_time"]), f"{st['pct']}%"])
    tbl3 = Table(lvl_data, colWidths=[6*cm, 6*cm, 6*cm])
    tbl3.setStyle(_tbl_style(GREEN))
    story.append(tbl3)
    story.append(Spacer(1, 0.4*cm))

    # ── Subject Analysis ─────────────────────────
    story.append(Paragraph("Performance by Subject", s["H2"]))
    subj_data = [["Subject", "Score %", "Remark"]]
    for subj, pct_s in result["SubjectStats"].items():
        remark = "✅ Strong" if pct_s >= 70 else ("⚠ Average" if pct_s >= 50 else "❌ Needs Work")
        subj_data.append([subj, f"{pct_s}%", remark])
    tbl4 = Table(subj_data, colWidths=[5*cm, 5*cm, 8*cm])
    tbl4.setStyle(_tbl_style(GOLD))
    story.append(tbl4)
    story.append(Spacer(1, 0.4*cm))

    # ── Recommendations ──────────────────────────
    story.append(Paragraph("Recommendations", s["H2"]))
    story.append(Paragraph(
        f"<b>🏆 Strongest Subject:</b>  {result['BestSubject']} "
        f"({result['SubjectStats'][result['BestSubject']]}%)", s["Body"]))
    story.append(Paragraph(
        f"<b>📚 Needs Improvement:</b>  {result['WorstSubject']} "
        f"({result['SubjectStats'][result['WorstSubject']]}%)", s["Body"]))
    story.append(Spacer(1, 0.3*cm))

    # ── Q-by-Q table ─────────────────────────────
    story.append(Paragraph("Question-wise Breakdown", s["H2"]))
    rec = result["Records"]
    q_data = [["Q#", "Subject", "Level", "Type", "Marks", "Max", "Time(s)", "✓"]]
    for _, row in rec.iterrows():
        tick = "✓" if row["Correct"] else "✗"
        q_data.append([
            int(row["Q_No"]), row["Subject"], row["Level"], row["Type"],
            row["Marks"], int(row["MaxMarks"]), int(row["Time"]), tick
        ])
    col_w = [1.2*cm, 3*cm, 3*cm, 1.8*cm, 2*cm, 2*cm, 2.5*cm, 1.5*cm]
    tbl5 = Table(q_data, colWidths=col_w, repeatRows=1)
    style5 = _tbl_style()
    # Colour correct/wrong column
    for i, row in enumerate(rec.itertuples(), start=1):
        bg = colors.HexColor("#C6EFCE") if row.Correct else colors.HexColor("#FFC7CE")
        style5.add("BACKGROUND", (7, i), (7, i), bg)
    tbl5.setStyle(style5)
    story.append(tbl5)

    doc.build(story)
    print(f"  [PDF] {os.path.basename(path)}")
    return path


# ─────────────────────────────────────────────
# STEP 4b – Examiner Summary PDF
# ─────────────────────────────────────────────
def generate_examiner_pdf(all_results, output_dir):
    s    = _styles()
    path = os.path.join(output_dir, "EXAMINER_Summary.pdf")
    doc  = SimpleDocTemplate(path, pagesize=A4,
                             leftMargin=1.8*cm, rightMargin=1.8*cm,
                             topMargin=1.5*cm, bottomMargin=1.5*cm)
    story = []

    story.append(Paragraph("📊 Examiner – Class Result Summary", s["H1"]))
    story.append(Paragraph(f"Total Students: {len(all_results)}", s["Sub"]))
    story.append(HRFlowable(width="100%", thickness=2, color=BRAND_BLUE))
    story.append(Spacer(1, 0.4*cm))

    # ── Class scoreboard ─────────────────────────
    story.append(Paragraph("Scoreboard", s["H2"]))
    board = [["Rank", "ID", "Name", "Scored", "Total", "%", "Grade"]]
    ranked = sorted(all_results, key=lambda r: r["Percentage"], reverse=True)
    for rank, r in enumerate(ranked, 1):
        board.append([rank, r["Student_ID"], r["Name"],
                      r["ScoredMarks"], r["TotalMarks"],
                      f"{r['Percentage']}%", r["Grade"].split("–")[0].strip()])
    tbl = Table(board, colWidths=[1.5*cm, 2*cm, 5.5*cm, 2.5*cm, 2.5*cm, 2.5*cm, 3*cm])
    tbl.setStyle(_tbl_style())
    story.append(tbl)
    story.append(Spacer(1, 0.4*cm))

    # ── Class statistics ─────────────────────────
    pcts = [r["Percentage"] for r in all_results]
    story.append(Paragraph("Class Statistics", s["H2"]))
    stat_data = [
        ["Metric", "Value"],
        ["Class Average",   f"{round(sum(pcts)/len(pcts), 2)}%"],
        ["Highest Score",   f"{max(pcts)}%"],
        ["Lowest Score",    f"{min(pcts)}%"],
        ["Pass Rate (≥40%)",f"{sum(1 for p in pcts if p >= 40)}/{len(pcts)}"],
    ]
    tbl2 = Table(stat_data, colWidths=[9*cm, 9*cm])
    tbl2.setStyle(_tbl_style(ACCENT))
    story.append(tbl2)
    story.append(Spacer(1, 0.4*cm))

    # ── Subject averages ─────────────────────────
    story.append(Paragraph("Average Score by Subject", s["H2"]))
    subj_avgs = {subj: round(sum(r["SubjectStats"][subj] for r in all_results) / len(all_results), 2)
                 for subj in SUBJECTS}
    subj_data = [["Subject", "Class Avg %"]]
    for subj, avg in sorted(subj_avgs.items(), key=lambda x: -x[1]):
        subj_data.append([subj, f"{avg}%"])
    tbl3 = Table(subj_data, colWidths=[9*cm, 9*cm])
    tbl3.setStyle(_tbl_style(GREEN))
    story.append(tbl3)
    story.append(Spacer(1, 0.4*cm))

    # ── Level averages ───────────────────────────
    story.append(Paragraph("Average Score by Difficulty Level", s["H2"]))
    lvl_data = [["Level", "Class Avg %", "Avg Time (sec)"]]
    for lvl in LEVELS:
        avg_pct  = round(sum(r["LevelStats"][lvl]["pct"]      for r in all_results) / len(all_results), 2)
        avg_time = round(sum(r["LevelStats"][lvl]["avg_time"] for r in all_results) / len(all_results), 1)
        lvl_data.append([lvl, f"{avg_pct}%", str(avg_time)])
    tbl4 = Table(lvl_data, colWidths=[6*cm, 6*cm, 6*cm])
    tbl4.setStyle(_tbl_style(GOLD))
    story.append(tbl4)
    story.append(Spacer(1, 0.4*cm))

    # ── Per student subject highlights ───────────
    story.append(Paragraph("Student Strengths & Improvement Areas", s["H2"]))
    hi_data = [["Name", "Best Subject", "Score%", "Needs Work", "Score%"]]
    for r in ranked:
        hi_data.append([
            r["Name"],
            r["BestSubject"],
            f"{r['SubjectStats'][r['BestSubject']]}%",
            r["WorstSubject"],
            f"{r['SubjectStats'][r['WorstSubject']]}%",
        ])
    tbl5 = Table(hi_data, colWidths=[4.5*cm, 3.5*cm, 2.5*cm, 3.5*cm, 2.5*cm])
    tbl5.setStyle(_tbl_style())
    story.append(tbl5)

    doc.build(story)
    print(f"  [PDF] {os.path.basename(path)}")
    return path


# ─────────────────────────────────────────────
# MAIN
# ─────────────────────────────────────────────
def main():
    print("\n=== Result Analysis System ===\n")

    print("Loading Question Bank...")
    qdf = pd.read_excel("question_bank_24_questions.xlsx")

    print("Loading Student Responses...")
    sdf = pd.read_excel("student_responses_10_students.xlsx")

    print("\nStep 3 – Analysing results …")
    all_results = []

    for _, srow in sdf.iterrows():
        res = analyse_student(srow, qdf)
        all_results.append(res)

    # Ranking
    ranked = sorted(
        all_results,
        key=lambda x: x["Percentage"],
        reverse=True
    )

    for rank, student in enumerate(ranked, start=1):
        student["Rank"] = rank

    print("\nStep 4 – Generating PDFs …")

    for res in ranked:
        generate_student_pdf(res, OUT_DIR)

    generate_examiner_pdf(ranked, OUT_DIR)

    print(f"\n✅ All files saved to: {OUT_DIR}")
    



if __name__ == "__main__":
    main()
